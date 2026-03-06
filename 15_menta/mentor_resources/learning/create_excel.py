# -*- coding: utf-8 -*-
"""
5カ年経営計画書テンプレート作成
PL/CF/BS の三表連動Excel
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# スタイル
hfont = Font(bold=True, size=12)
tfont = Font(bold=True, size=14)
input_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
calc_fill = PatternFill(start_color='E0F0E0', end_color='E0F0E0', fill_type='solid')
check_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
yen = '#,##0'

def sc(ws, row, col, value=None, font=None, fill=None, fmt=None, formula=None):
    cell = ws.cell(row=row, column=col)
    if formula:
        cell.value = formula
    elif value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    cell.border = border
    return cell

years = ['R2', 'R3', 'R4', 'R5', 'R6']

# ========== PL ==========
ws = wb.active
ws.title = 'PL'
ws.column_dimensions['A'].width = 25
for c in ['B','C','D','E','F']:
    ws.column_dimensions[c].width = 18

sc(ws, 1, 1, 'PL', tfont)
for i, y in enumerate(years):
    sc(ws, 2, i+2, y, hfont)

# 行番号
R = {'train':3, 'welf':4, 'other':5, 'rev':6,
     'cost':8, 'gross':9, 'exp':10, 'pers':11, 'oper':12,
     'ext':14, 'int':15, 'ord':16, 'tax':17, 'net':18}

labels = {3:'訓練給付金等', 4:'福祉的就労売上', 5:'その他収入', 6:'総売上',
          7:'', 8:'事業原価', 9:'売上総利益', 10:'経費', 11:'総人件費（参考）', 12:'営業利益',
          13:'', 14:'営業外収益', 15:'支払利息', 16:'経常利益', 17:'法人税等（40%）', 18:'税引後利益'}

for r, lb in labels.items():
    bold = r in [6,9,12,16,18]
    sc(ws, r, 1, lb, hfont if bold else None)

# 入力データ
data = {
    'R2': [43054612, 9210000, 0, 49636638, 9566638, 43375000, 245000, 0],
    'R3': [114485450, 22380025, 0, 104757564, 19804564, 87533000, 420000, 0],
    'R4': [139618300, 26820027, 0, 81972549, 18180549, 104856000, 420000, 0],
}
input_keys = ['train','welf','other','cost','exp','pers','int','ext']

for i, y in enumerate(years):
    col = i + 2
    c = get_column_letter(col)

    # 入力セル
    vals = data.get(y, [0]*8)
    for j, key in enumerate(input_keys):
        sc(ws, R[key], col, vals[j], fill=input_fill, fmt=yen)

    # 数式セル
    sc(ws, R['rev'], col, fill=calc_fill, fmt=yen,
       formula='={0}{1}+{0}{2}+{0}{3}'.format(c, R['train'], R['welf'], R['other']))
    sc(ws, R['gross'], col, fill=calc_fill, fmt=yen,
       formula='={0}{1}-{0}{2}'.format(c, R['rev'], R['cost']))
    sc(ws, R['oper'], col, fill=calc_fill, fmt=yen,
       formula='={0}{1}-{0}{2}'.format(c, R['gross'], R['exp']))
    sc(ws, R['ord'], col, fill=calc_fill, fmt=yen,
       formula='={0}{1}+{0}{2}-{0}{3}'.format(c, R['oper'], R['ext'], R['int']))
    sc(ws, R['tax'], col, fill=calc_fill, fmt=yen,
       formula='=MAX({0}{1}*0.4,0)'.format(c, R['ord']))
    sc(ws, R['net'], col, fill=calc_fill, fmt=yen,
       formula='={0}{1}-{0}{2}'.format(c, R['ord'], R['tax']))

# ========== CF ==========
ws2 = wb.create_sheet('CF')
ws2.column_dimensions['A'].width = 25
for c in ['B','C','D','E','F']:
    ws2.column_dimensions[c].width = 18

sc(ws2, 1, 1, 'CF', tfont)
for i, y in enumerate(years):
    sc(ws2, 2, i+2, y, hfont)

C = {'open':3, 'opi':5, 'ope':6, 'opcf':7, 'fi':9, 'fe':10, 'fcf':11, 'tcf':13, 'tax':14, 'close':15}

cf_labels = {3:'前期繰越', 4:'【経常収支】', 5:'経常収入', 6:'経常支出', 7:'経常収支差額',
             8:'【財務収支】', 9:'財務収入', 10:'財務支出', 11:'財務収支差額',
             12:'', 13:'CF合計', 14:'法人税等', 15:'期末残高'}

for r, lb in cf_labels.items():
    bold = r in [7,11,13,15]
    sc(ws2, r, 1, lb, hfont if bold else None)

cf_data = {
    'R2': {'open':0, 'opi':40560360, 'ope':59863907, 'fi':23000000, 'fe':3399000},
    'R3': {'opi':129071704, 'ope':128400542, 'fi':0, 'fe':684000},
    'R4': {'opi':146958399, 'ope':137554977, 'fi':0, 'fe':684000},
}

for i, y in enumerate(years):
    col = i + 2
    c = get_column_letter(col)

    # 前期繰越
    if y == 'R2':
        sc(ws2, C['open'], col, 0, fill=input_fill, fmt=yen)
    else:
        pc = get_column_letter(col-1)
        sc(ws2, C['open'], col, fill=calc_fill, fmt=yen,
           formula='={0}{1}'.format(pc, C['close']))

    # 入力
    d = cf_data.get(y, {})
    for key in ['opi','ope','fi','fe']:
        sc(ws2, C[key], col, d.get(key, 0), fill=input_fill, fmt=yen)

    # 数式
    sc(ws2, C['opcf'], col, fill=calc_fill, fmt=yen,
       formula='={0}{1}-{0}{2}'.format(c, C['opi'], C['ope']))
    sc(ws2, C['fcf'], col, fill=calc_fill, fmt=yen,
       formula='={0}{1}-{0}{2}'.format(c, C['fi'], C['fe']))
    sc(ws2, C['tcf'], col, fill=calc_fill, fmt=yen,
       formula='={0}{1}+{0}{2}'.format(c, C['opcf'], C['fcf']))
    sc(ws2, C['tax'], col, fill=calc_fill, fmt=yen,
       formula="=PL!{0}{1}".format(c, R['tax']))
    sc(ws2, C['close'], col, fill=calc_fill, fmt=yen,
       formula='={0}{1}+{0}{2}-{0}{3}'.format(c, C['open'], C['tcf'], C['tax']))

# ========== BS ==========
ws3 = wb.create_sheet('BS')
ws3.column_dimensions['A'].width = 25
for c in ['B','C','D','E','F']:
    ws3.column_dimensions[c].width = 18

sc(ws3, 1, 1, 'BS', tfont)
for i, y in enumerate(years):
    sc(ws3, 2, i+2, y, hfont)

bs_labels = {3:'【資産】', 4:'現金預金', 5:'売掛金', 6:'流動資産合計', 7:'固定資産', 8:'資産合計',
             9:'', 10:'【負債】', 11:'買掛金', 12:'流動負債合計', 13:'長期借入金', 14:'短期借入金',
             15:'固定負債合計', 16:'負債合計',
             17:'', 18:'【純資産】', 19:'資本金', 20:'利益剰余金', 21:'純資産合計',
             22:'', 23:'負債+純資産合計', 24:'', 25:'貸借チェック（0=OK）'}

for r, lb in bs_labels.items():
    bold = r in [6,8,12,15,16,21,23,25]
    sc(ws3, r, 1, lb, hfont if bold else None)

bs_data = {
    'R2': {'recv':12504252, 'fixed':0, 'pay':7122731, 'llong':14601000, 'lshort':8000000, 'cap':0},
    'R3': {'recv':22754103, 'fixed':0, 'pay':9919753, 'llong':13917000, 'lshort':8000000, 'cap':0},
    'R4': {'recv':0, 'fixed':0, 'pay':0, 'llong':13290000, 'lshort':8000000, 'cap':0},
}
bs_map = {'recv':5, 'fixed':7, 'pay':11, 'llong':13, 'lshort':14, 'cap':19}

for i, y in enumerate(years):
    col = i + 2
    c = get_column_letter(col)

    # 現金 = CF期末残高
    sc(ws3, 4, col, fill=calc_fill, fmt=yen,
       formula="=CF!{0}{1}".format(c, C['close']))

    # 入力セル
    d = bs_data.get(y, {'recv':0,'fixed':0,'pay':0,'llong':0,'lshort':0,'cap':0})
    for key, row in bs_map.items():
        sc(ws3, row, col, d[key], fill=input_fill, fmt=yen)

    # 数式
    sc(ws3, 6, col, fill=calc_fill, fmt=yen, formula='={0}4+{0}5'.format(c))
    sc(ws3, 8, col, fill=calc_fill, fmt=yen, formula='={0}6+{0}7'.format(c))
    sc(ws3, 12, col, fill=calc_fill, fmt=yen, formula='={0}11'.format(c))
    sc(ws3, 15, col, fill=calc_fill, fmt=yen, formula='={0}13+{0}14'.format(c))
    sc(ws3, 16, col, fill=calc_fill, fmt=yen, formula='={0}12+{0}15'.format(c))

    # 利益剰余金 = 前年 + 当期純利益
    if y == 'R2':
        sc(ws3, 20, col, fill=calc_fill, fmt=yen,
           formula="=PL!{0}{1}".format(c, R['net']))
    else:
        pc = get_column_letter(col-1)
        sc(ws3, 20, col, fill=calc_fill, fmt=yen,
           formula='={0}20+PL!{1}{2}'.format(pc, c, R['net']))

    sc(ws3, 21, col, fill=calc_fill, fmt=yen, formula='={0}19+{0}20'.format(c))
    sc(ws3, 23, col, fill=calc_fill, fmt=yen, formula='={0}16+{0}21'.format(c))
    sc(ws3, 25, col, fill=check_fill, fmt=yen, formula='={0}8-{0}23'.format(c))

# ========== 凡例 ==========
ws4 = wb.create_sheet('凡例')
ws4.column_dimensions['A'].width = 50
ws4['A1'] = '色の意味'
ws4['A1'].font = tfont
sc(ws4, 3, 1, '黄色 = 入力セル（手で入力する箇所）', fill=input_fill)
sc(ws4, 4, 1, '緑色 = 自動計算（数式で連動）', fill=calc_fill)
sc(ws4, 5, 1, '赤色 = チェック（0なら貸借一致）', fill=check_fill)
ws4['A7'] = '三表連動ルール'
ws4['A7'].font = tfont
ws4['A8'] = 'PL税引後利益 → BS利益剰余金（累積加算）'
ws4['A9'] = 'CF期末残高 → BS現金預金'
ws4['A10'] = 'PL法人税 → CF法人税等'
ws4['A11'] = 'CF前期繰越 = 前年CF期末残高'
ws4['A12'] = 'BS貸借チェック = 資産合計 - (負債+純資産) → 0で一致'

out = 'C:/Users/USER/Documents/MENTA/mentor_resources/learning/5year_plan_template.xlsx'
wb.save(out)
print('Created:', out)
